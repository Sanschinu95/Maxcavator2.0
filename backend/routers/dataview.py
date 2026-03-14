"""
routers/dataview.py — Explore endpoints

GET /explore/{doc_id}/sections
GET /explore/{doc_id}/tables
GET /explore/{doc_id}/images
GET /explore/{doc_id}/links
GET /explore/{doc_id}/json
GET /explore/{doc_id}/images/{image_index}/download
GET /explore/{doc_id}/tables/{table_index}/export?format=csv|json|xlsx
GET /explore/{doc_id}/sections/{section_index}/download
GET /explore/{doc_id}/links/export
"""

import csv
import io
import json as json_module
from pathlib import Path

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import FileResponse, Response, StreamingResponse
from database import async_get_document
from config import IMAGES_DIR

router = APIRouter(prefix="/explore", tags=["explore"])


async def _get_doc_or_404(doc_id: str) -> dict:
    doc = await async_get_document(doc_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found.")
    return doc


@router.get("/{doc_id}/sections")
async def get_sections(doc_id: str):
    doc = await _get_doc_or_404(doc_id)
    return {
        "doc_id":   doc_id,
        "title":    doc.get("metadata", {}).get("title", doc.get("filename", "")),
        "sections": doc.get("sections", []),
    }


@router.get("/{doc_id}/tables")
async def get_tables(doc_id: str):
    doc = await _get_doc_or_404(doc_id)
    return {
        "doc_id": doc_id,
        "tables": doc.get("tables", []),
    }


@router.get("/{doc_id}/images")
async def get_images(doc_id: str):
    doc = await _get_doc_or_404(doc_id)
    return {
        "doc_id": doc_id,
        "images": doc.get("images", []),
    }


@router.get("/{doc_id}/links")
async def get_links(doc_id: str):
    doc = await _get_doc_or_404(doc_id)
    return {
        "doc_id": doc_id,
        "links":  doc.get("links", []),
    }


@router.get("/{doc_id}/json")
async def get_full_json(doc_id: str):
    doc = await _get_doc_or_404(doc_id)
    return doc


# ─────────────────────────────────────────────────────────────────────────────
# UPGRADE 1 + 2: Download / Export endpoints
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/{doc_id}/images/{image_index}/download")
async def download_image(doc_id: str, image_index: int):
    """Download an extracted image as PNG."""
    doc = await _get_doc_or_404(doc_id)
    images = doc.get("images", [])

    # Find the image entry matching this index
    img_entry = None
    for img in images:
        if img.get("image_index") == image_index:
            img_entry = img
            break
    # Also try by list position if image_index field doesn't match
    if img_entry is None and 0 <= image_index < len(images):
        img_entry = images[image_index]

    if img_entry is None:
        raise HTTPException(status_code=404, detail="Image not found.")

    # Resolve file path from image_path (e.g. "images/{doc_id}/page1_img0.png")
    image_path = img_entry.get("image_path", "")
    filename = image_path.split("/")[-1] if "/" in image_path else image_path

    file_path = IMAGES_DIR / doc_id / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Image file not found on disk.")

    return FileResponse(
        path=str(file_path),
        media_type="image/png",
        filename=filename,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{doc_id}/sections/{section_index}/download")
async def download_section(doc_id: str, section_index: int):
    """Download a section as a .txt file."""
    doc = await _get_doc_or_404(doc_id)
    sections = doc.get("sections", [])

    if section_index < 0 or section_index >= len(sections):
        raise HTTPException(status_code=404, detail="Section not found.")

    section = sections[section_index]
    heading = section.get("heading", f"Section {section_index + 1}")
    content = section.get("content", "")
    text_content = f"{heading}\n\n{content}"

    safe_heading = "".join(c if c.isalnum() or c in " -_" else "" for c in heading)[:50].strip()
    filename = f"{safe_heading or f'section_{section_index + 1}'}.txt"

    return Response(
        content=text_content,
        media_type="text/plain; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{doc_id}/links/export")
async def export_links_csv(doc_id: str):
    """Export all links as a CSV file."""
    doc = await _get_doc_or_404(doc_id)
    links = doc.get("links", [])

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["text", "url", "page"])
    for link in links:
        writer.writerow([
            link.get("text", ""),
            link.get("url", ""),
            link.get("page", ""),
        ])

    csv_bytes = output.getvalue().encode("utf-8")
    return Response(
        content=csv_bytes,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{doc_id}_links.csv"'},
    )


@router.get("/{doc_id}/tables/{table_index}/export")
async def export_table(doc_id: str, table_index: int, format: str = Query("csv")):
    """Export a single table as CSV, JSON, or XLSX."""
    doc = await _get_doc_or_404(doc_id)
    tables = doc.get("tables", [])

    if table_index < 0 or table_index >= len(tables):
        raise HTTPException(status_code=404, detail="Table not found.")

    table = tables[table_index]
    headers = table.get("headers", [])
    rows = table.get("rows", [])
    fmt = format.lower().strip()

    if fmt == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        csv_bytes = output.getvalue().encode("utf-8")
        return Response(
            content=csv_bytes,
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="table_{table_index + 1}.csv"'},
        )

    elif fmt == "json":
        payload = json_module.dumps({"headers": headers, "rows": rows}, indent=2, ensure_ascii=False)
        return Response(
            content=payload.encode("utf-8"),
            media_type="application/json",
            headers={"Content-Disposition": f'attachment; filename="table_{table_index + 1}.json"'},
        )

    elif fmt == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = f"Table {table_index + 1}"

        # Header row in bold
        bold_font = Font(bold=True)
        for ci, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=ci, value=header)
            cell.font = bold_font

        # Data rows
        for ri, row in enumerate(rows, start=2):
            for ci, cell_val in enumerate(row, start=1):
                ws.cell(row=ri, column=ci, value=cell_val)

        # Auto-width columns
        for ci, header in enumerate(headers, start=1):
            max_len = len(str(header))
            for ri, row in enumerate(rows):
                if ci - 1 < len(row):
                    max_len = max(max_len, len(str(row[ci - 1])))
            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max_len + 4, 50)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="table_{table_index + 1}.xlsx"'},
        )

    else:
        raise HTTPException(status_code=400, detail="Format must be csv, json, or xlsx.")
